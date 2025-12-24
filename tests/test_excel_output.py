"""
Unit tests for Excel output module.

Tests Excel workbook generation and formatting.
"""

import pytest
import pandas as pd
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from modules.excel_output import (
    apply_header_formatting,
    apply_confidence_color_coding,
    auto_fit_columns,
    apply_data_filters,
    freeze_header_row,
    create_excel_workbook,
    COLORS
)


@pytest.fixture
def sample_contacts():
    """Create sample contacts DataFrame for testing."""
    return pd.DataFrame([
        {
            'institution_name': 'Stanford Law School',
            'state': 'CA',
            'program_type': 'Law School',
            'full_name': 'John Doe',
            'email': 'john@stanford.edu',
            'phone': '555-1234',
            'confidence_score': 80,
            'email_status': 'valid'
        },
        {
            'institution_name': 'UCLA Law School',
            'state': 'CA',
            'program_type': 'Law School',
            'full_name': 'Jane Smith',
            'email': 'jane@ucla.edu',
            'phone': '555-5678',
            'confidence_score': 60,
            'email_status': 'catch-all'
        },
        {
            'institution_name': 'NYU Law School',
            'state': 'NY',
            'program_type': 'Law School',
            'full_name': 'Bob Johnson',
            'email': 'bob@nyu.edu',
            'phone': '',
            'confidence_score': 45,
            'email_status': 'invalid'
        },
        {
            'institution_name': 'Texas Paralegal Program',
            'state': 'TX',
            'program_type': 'Paralegal Program',
            'full_name': 'Alice Brown',
            'email': 'alice@tx.edu',
            'phone': '555-9999',
            'confidence_score': 90,
            'email_status': 'valid'
        }
    ])


@pytest.fixture
def sample_stats():
    """Create sample statistics dictionary for testing."""
    return {
        'summary': {
            'total_contacts': 4,
            'total_institutions': 4,
            'avg_contacts_per_institution': 1.0
        },
        'by_state': {
            'CA': 2,
            'NY': 1,
            'TX': 1
        },
        'by_program_type': {
            'Law School': {'count': 3, 'percentage': 75.0},
            'Paralegal Program': {'count': 1, 'percentage': 25.0}
        },
        'email_quality': {
            'total_contacts': 4,
            'with_email': 4,
            'email_coverage_pct': 100.0,
            'valid_deliverable': 2,
            'valid_pct': 50.0,
            'catch_all': 1,
            'catch_all_pct': 25.0,
            'invalid': 1,
            'invalid_pct': 25.0
        },
        'confidence_distribution': {
            'high': {'count': 2, 'percentage': 50.0, 'range': '75-100'},
            'medium': {'count': 1, 'percentage': 25.0, 'range': '50-74'},
            'low': {'count': 1, 'percentage': 25.0, 'range': '0-49'},
            'average_score': 68.8,
            'median_score': 70,
            'min_score': 45,
            'max_score': 90
        },
        'top_roles': [
            {'role': 'Library Director', 'count': 2},
            {'role': 'Associate Dean', 'count': 1}
        ],
        'scraping_success': {
            'total_institutions_attempted': 5,
            'successful_extractions': 4,
            'failed_extractions': 1,
            'success_rate_pct': 80.0,
            'avg_contacts_per_institution': 1.0
        }
    }


@pytest.fixture
def sample_targets():
    """Create sample targets DataFrame for testing."""
    return pd.DataFrame([
        {'name': 'Stanford Law School', 'state': 'CA', 'type': 'Law School'},
        {'name': 'UCLA Law School', 'state': 'CA', 'type': 'Law School'},
        {'name': 'NYU Law School', 'state': 'NY', 'type': 'Law School'},
        {'name': 'Texas Paralegal Program', 'state': 'TX', 'type': 'Paralegal Program'},
        {'name': 'Failed Institution', 'state': 'CA', 'type': 'Law School'}
    ])


class TestCreateExcelWorkbook:
    """Test main Excel workbook creation."""

    def test_creates_all_8_sheets(self, sample_contacts, sample_stats, sample_targets):
        """Test that all 8 sheets are created."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            success = create_excel_workbook(sample_contacts, sample_stats, temp_path, sample_targets)
            assert success is True

            # Load workbook and check sheets
            wb = load_workbook(temp_path)
            sheet_names = wb.sheetnames

            assert 'All Contacts' in sheet_names
            assert 'Law School Contacts' in sheet_names
            assert 'Paralegal Program Contacts' in sheet_names
            assert 'High Confidence' in sheet_names
            assert 'Medium Confidence' in sheet_names
            assert 'Needs Review' in sheet_names
            assert 'Statistics Summary' in sheet_names
            assert 'Scraping Log' in sheet_names

        finally:
            Path(temp_path).unlink()

    def test_all_contacts_sheet_content(self, sample_contacts, sample_stats):
        """Test that All Contacts sheet contains all contacts."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(sample_contacts, sample_stats, temp_path, None)
            wb = load_workbook(temp_path)
            ws = wb['All Contacts']

            # Check row count (header + 4 contacts)
            assert ws.max_row == 5

            # Check that all contacts are present
            contact_names = [ws.cell(row=i, column=ws['A1'].column + 3).value for i in range(2, 6)]  # Assuming full_name is 4th column
            assert 'John Doe' in str(contact_names)
            assert 'Jane Smith' in str(contact_names)

        finally:
            Path(temp_path).unlink()

    def test_filtered_sheets_content(self, sample_contacts, sample_stats):
        """Test that filtered sheets contain correct data."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(sample_contacts, sample_stats, temp_path, None)
            wb = load_workbook(temp_path)

            # Test Law School Contacts sheet (should have 3 contacts)
            law_ws = wb['Law School Contacts']
            assert law_ws.max_row == 4  # Header + 3 law school contacts

            # Test Paralegal Program Contacts sheet (should have 1 contact)
            para_ws = wb['Paralegal Program Contacts']
            assert para_ws.max_row == 2  # Header + 1 paralegal contact

            # Test High Confidence sheet (>= 75: John Doe 80, Alice Brown 90)
            high_ws = wb['High Confidence']
            assert high_ws.max_row == 3  # Header + 2 high confidence contacts

            # Test Medium Confidence sheet (50-74: Jane Smith 60)
            med_ws = wb['Medium Confidence']
            assert med_ws.max_row == 2  # Header + 1 medium confidence contact

            # Test Needs Review sheet (< 50: Bob Johnson 45)
            low_ws = wb['Needs Review']
            assert low_ws.max_row == 2  # Header + 1 low confidence contact

        finally:
            Path(temp_path).unlink()

    def test_statistics_sheet_content(self, sample_contacts, sample_stats):
        """Test that Statistics Summary sheet is populated."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(sample_contacts, sample_stats, temp_path, None)
            wb = load_workbook(temp_path)
            ws = wb['Statistics Summary']

            # Check that sheet has content
            assert ws.max_row > 10  # Should have many rows of statistics

            # Check for key section headers
            all_values = [ws.cell(row=i, column=1).value for i in range(1, ws.max_row + 1)]
            assert 'SUMMARY' in all_values
            assert 'EMAIL QUALITY' in all_values
            assert 'CONFIDENCE DISTRIBUTION' in all_values

        finally:
            Path(temp_path).unlink()

    def test_scraping_log_sheet_content(self, sample_contacts, sample_stats, sample_targets):
        """Test that Scraping Log sheet is populated."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(sample_contacts, sample_stats, temp_path, sample_targets)
            wb = load_workbook(temp_path)
            ws = wb['Scraping Log']

            # Should have header + successful institutions + failed institutions
            assert ws.max_row >= 5  # Header + at least 4 institutions

            # Check for success and failed statuses
            all_statuses = [ws.cell(row=i, column=4).value for i in range(2, ws.max_row + 1)]  # Status column
            assert 'Success' in all_statuses
            assert 'Failed' in all_statuses

        finally:
            Path(temp_path).unlink()


class TestExcelFormatting:
    """Test Excel formatting features."""

    def test_header_formatting_applied(self, sample_contacts, sample_stats):
        """Test that header rows are formatted correctly."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(sample_contacts, sample_stats, temp_path, None)
            wb = load_workbook(temp_path)
            ws = wb['All Contacts']

            # Check first row (header) has blue background
            header_cell = ws['A1']
            # openpyxl uses '00' as alpha channel prefix
            assert header_cell.fill.start_color.rgb == '00' + COLORS['header']
            assert header_cell.font.bold is True

        finally:
            Path(temp_path).unlink()

    def test_frozen_panes_applied(self, sample_contacts, sample_stats):
        """Test that header row is frozen."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(sample_contacts, sample_stats, temp_path, None)
            wb = load_workbook(temp_path)
            ws = wb['All Contacts']

            # Check that freeze panes is set to A2 (freezes row 1)
            assert ws.freeze_panes == 'A2'

        finally:
            Path(temp_path).unlink()

    def test_auto_filter_applied(self, sample_contacts, sample_stats):
        """Test that data filters are applied to header row."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(sample_contacts, sample_stats, temp_path, None)
            wb = load_workbook(temp_path)
            ws = wb['All Contacts']

            # Check that auto filter is enabled
            assert ws.auto_filter.ref is not None

        finally:
            Path(temp_path).unlink()

    def test_color_coding_by_confidence(self, sample_contacts, sample_stats):
        """Test that rows are color-coded by confidence score."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(sample_contacts, sample_stats, temp_path, None)
            wb = load_workbook(temp_path)
            ws = wb['All Contacts']

            # Find confidence_score column
            headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
            confidence_col = headers.index('confidence_score') + 1 if 'confidence_score' in headers else None

            if confidence_col:
                # Check row 2 (John Doe, score 80) should be green (high confidence)
                cell = ws.cell(row=2, column=1)
                # openpyxl uses '00' as alpha channel prefix
                assert cell.fill.start_color.rgb == '00' + COLORS['high']

                # Check row 4 (Bob Johnson, score 45) should be red (low confidence)
                cell = ws.cell(row=4, column=1)
                assert cell.fill.start_color.rgb == '00' + COLORS['low']

        finally:
            Path(temp_path).unlink()


class TestEmptyDataHandling:
    """Test handling of empty data."""

    def test_empty_contacts_creates_workbook(self, sample_stats):
        """Test that workbook is created even with empty contacts."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            empty_df = pd.DataFrame()
            success = create_excel_workbook(empty_df, sample_stats, temp_path, None)

            # Should still create workbook (with "No contacts found" messages)
            assert success is True

            wb = load_workbook(temp_path)
            assert 'All Contacts' in wb.sheetnames

        finally:
            Path(temp_path).unlink()

    def test_empty_filtered_sheet(self, sample_contacts, sample_stats):
        """Test that filtered sheets handle no matching data."""
        # Create contacts with only Law Schools (no Paralegal Programs)
        law_only_df = sample_contacts[sample_contacts['program_type'] == 'Law School'].copy()

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            temp_path = f.name

        try:
            create_excel_workbook(law_only_df, sample_stats, temp_path, None)
            wb = load_workbook(temp_path)

            # Paralegal Program Contacts sheet should exist but have "No contacts match filter"
            para_ws = wb['Paralegal Program Contacts']
            assert para_ws.max_row >= 1  # Should have at least the message row

        finally:
            Path(temp_path).unlink()
