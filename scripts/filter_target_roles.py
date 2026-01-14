#!/usr/bin/env python3
"""
Role-Based Contact Filtering System

Filters 15,603 contacts to identify high-priority targets matching 10 specific roles
across Law Schools and Paralegal Programs using intelligent fuzzy matching.

Author: Generated for Legal Education Contact Scraper
Date: 2026-01-09
"""

import pandas as pd
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict
from datetime import datetime
import re


class RoleMatchingEngine:
    """Intelligent role matching with fuzzy matching, synonyms, and context awareness"""

    # Define target roles with variations and synonyms
    LAW_SCHOOL_ROLES = {
        "Law Library Director": {
            "primary": "Law Library Director",
            "synonyms": [
                "Library Director",
                "Head Librarian",
                "Law Librarian",
                "Director of Library Services",
                "Director of the Law Library",
                "Associate Director Law Library",
                "Assistant Director Law Library"
            ],
            "keywords": ["library", "librarian"],
            "context_required": ["director", "head", "associate director", "assistant director"]
        },
        "Associate Dean for Academic Affairs": {
            "primary": "Associate Dean for Academic Affairs",
            "synonyms": [
                "Associate Dean Academic",
                "Assistant Dean Academic Affairs",
                "Academic Dean",
                "Associate Dean of Academic Affairs",
                "Assistant Dean for Academic Affairs",
                "Vice Dean Academic Affairs",
                "Vice Dean for Academic Affairs"
            ],
            "keywords": ["dean", "academic"],
            "context_required": ["associate", "assistant", "vice"]
        },
        "Legal Writing Director": {
            "primary": "Legal Writing Director",
            "synonyms": [
                "Director of Legal Writing",
                "Legal Writing Program Director",
                "Director Legal Research and Writing",
                "Director of Legal Research & Writing",
                "Director of the Legal Writing Program",
                "Legal Research and Writing Director",
                "Assistant Director Legal Writing"
            ],
            "keywords": ["legal writing", "legal research"],
            "context_required": ["director", "program director"]
        },
        "Experiential Learning Director": {
            "primary": "Experiential Learning Director",
            "synonyms": [
                "Clinical Director",
                "Externship Director",
                "Skills Director",
                "Practical Training Director",
                "Director of Experiential Learning",
                "Director of Clinical Programs",
                "Director of Externships",
                "Associate Dean Experiential Learning",
                "Assistant Dean Clinical Programs"
            ],
            "keywords": ["experiential", "clinical", "externship", "skills", "practical"],
            "context_required": ["director", "dean"]
        },
        "Instructional Technology Librarian": {
            "primary": "Instructional Technology Librarian",
            "synonyms": [
                "Technology Librarian",
                "Digital Services Librarian",
                "IT Librarian",
                "Instructional Technologist",
                "Educational Technology Librarian",
                "Technology Services Librarian"
            ],
            "keywords": ["technology", "librarian", "digital"],
            "context_required": ["instructional", "educational", "it", "digital"]
        }
    }

    PARALEGAL_ROLES = {
        "Paralegal Program Director": {
            "primary": "Paralegal Program Director",
            "synonyms": [
                "Director Paralegal Studies",
                "Paralegal Coordinator",
                "Paralegal Program Coordinator",
                "Director of Paralegal Program",
                "Paralegal Studies Director",
                "Program Director Paralegal",
                "Coordinator Paralegal Studies"
            ],
            "keywords": ["paralegal"],
            "context_required": ["director", "coordinator", "program"]
        },
        "Dean of Workforce Programs": {
            "primary": "Dean of Workforce Programs",
            "synonyms": [
                "Workforce Development Dean",
                "Career Programs Dean",
                "Dean Workforce Development",
                "Dean of Workforce Development",
                "Dean Career and Technical Education",
                "Dean CTE Programs"
            ],
            "keywords": ["workforce", "dean"],
            "context_required": ["development", "programs", "career"]
        },
        "Legal Studies Faculty": {
            "primary": "Legal Studies Faculty",
            "synonyms": [
                "Paralegal Instructor",
                "Legal Studies Professor",
                "Paralegal Studies Faculty",
                "Legal Studies Instructor",
                "Paralegal Professor",
                "Paralegal Studies Instructor"
            ],
            "keywords": ["legal studies", "paralegal"],
            "context_required": ["faculty", "instructor", "professor", "teacher"]
        },
        "Program Chair": {
            "primary": "Program Chair",
            "synonyms": [
                "Department Chair",
                "Program Coordinator",
                "Division Chair",
                "Chair Paralegal",
                "Chair Legal Studies",
                "Department Head"
            ],
            "keywords": ["chair", "coordinator"],
            "context_required": ["program", "department", "division"]
        },
        "Academic Affairs": {
            "primary": "Academic Affairs",
            "synonyms": [
                "Director Academic Affairs",
                "Associate Dean Academic Affairs",
                "Academic Services Director",
                "Director of Academic Services",
                "Academic Success Director",
                "Dean of Academic Affairs"
            ],
            "keywords": ["academic"],
            "context_required": ["affairs", "services", "success", "director", "dean"]
        }
    }

    def __init__(self, min_score=65):
        """Initialize with minimum match score threshold"""
        self.min_score = min_score
        self.all_roles = {**self.LAW_SCHOOL_ROLES, **self.PARALEGAL_ROLES}

    def normalize_title(self, title):
        """Normalize title for better matching"""
        if pd.isna(title) or not title:
            return ""

        # Convert to lowercase
        title = str(title).lower()

        # Remove common noise
        title = re.sub(r'\s+', ' ', title)  # Multiple spaces to single
        title = re.sub(r'[^\w\s&/-]', '', title)  # Remove special chars except &, /, -

        return title.strip()

    def match_role(self, title, program_type):
        """
        Match a title against target roles using multi-layered approach

        Returns: (best_role, match_score, match_type, match_explanation)
        """
        if not title or pd.isna(title):
            return None, 0, None, None

        normalized_title = self.normalize_title(title)

        # Select appropriate role set
        if program_type == "Law School":
            roles_to_check = self.LAW_SCHOOL_ROLES
        else:
            roles_to_check = self.PARALEGAL_ROLES

        best_match = None
        best_score = 0
        best_type = None
        best_explanation = None

        for role_name, role_config in roles_to_check.items():
            # Try exact match on primary role
            score, match_type, explanation = self._calculate_match_score(
                normalized_title, role_name, role_config
            )

            if score > best_score:
                best_score = score
                best_match = role_name
                best_type = match_type
                best_explanation = explanation

        # Only return if meets minimum score
        if best_score >= self.min_score:
            return best_match, best_score, best_type, best_explanation

        return None, 0, None, None

    def _calculate_match_score(self, normalized_title, role_name, role_config):
        """Calculate match score using multiple strategies"""

        primary = self.normalize_title(role_config['primary'])
        synonyms = [self.normalize_title(s) for s in role_config['synonyms']]
        keywords = [self.normalize_title(k) for k in role_config['keywords']]
        context_required = [self.normalize_title(c) for c in role_config['context_required']]

        # Strategy 1: Exact match (100 score)
        if normalized_title == primary:
            return 100, "Exact", f"Exact match: '{role_config['primary']}'"

        # Strategy 2: Exact synonym match (95 score)
        for synonym in synonyms:
            if normalized_title == synonym:
                return 95, "Exact Synonym", f"Exact synonym: '{synonym}'"

        # Strategy 3: Fuzzy match on primary (70-90 score)
        fuzzy_score = fuzz.token_set_ratio(normalized_title, primary)
        if fuzzy_score >= 90:
            return fuzzy_score, "Fuzzy Primary", f"Strong fuzzy match: {fuzzy_score}% similar to '{role_config['primary']}'"

        # Strategy 4: Fuzzy match on synonyms (70-88 score)
        best_synonym_score = 0
        best_synonym = None
        for synonym in synonyms:
            score = fuzz.token_set_ratio(normalized_title, synonym)
            if score > best_synonym_score:
                best_synonym_score = score
                best_synonym = synonym

        if best_synonym_score >= 85:
            return best_synonym_score - 2, "Fuzzy Synonym", f"Fuzzy synonym match: {best_synonym_score}% similar to '{best_synonym}'"

        # Strategy 5: Context-aware keyword matching (65-80 score)
        # Must contain at least one keyword AND one context word
        has_keyword = any(kw in normalized_title for kw in keywords)
        has_context = any(ctx in normalized_title for ctx in context_required)

        if has_keyword and has_context:
            # Calculate score based on how many keywords/context words match
            keyword_count = sum(1 for kw in keywords if kw in normalized_title)
            context_count = sum(1 for ctx in context_required if ctx in normalized_title)

            # Base score + bonuses
            context_score = 65 + (keyword_count * 5) + (context_count * 3)
            context_score = min(context_score, 80)  # Cap at 80

            matched_keywords = [kw for kw in keywords if kw in normalized_title]
            matched_context = [ctx for ctx in context_required if ctx in normalized_title]

            return context_score, "Context", f"Context match: keywords={matched_keywords}, context={matched_context}"

        # No match
        return 0, None, None


def filter_contacts_by_roles(input_file, output_file, min_score=65):
    """Main filtering function"""

    print("=" * 80)
    print("ROLE-BASED CONTACT FILTERING SYSTEM")
    print("=" * 80)
    print(f"\nInput file: {input_file}")
    print(f"Minimum match score: {min_score}")
    print(f"Output file: {output_file}")

    # Load data
    print("\n[1/6] Loading contact database...")
    df = pd.read_excel(input_file, sheet_name='Master Contact List')
    print(f"  Total contacts: {len(df):,}")
    print(f"  Law Schools: {(df['Organization Type'] == 'Law School').sum():,}")
    print(f"  Paralegal Programs: {(df['Organization Type'] == 'Paralegal Program').sum():,}")

    # Initialize matching engine
    print(f"\n[2/6] Initializing role matching engine (min score: {min_score})...")
    matcher = RoleMatchingEngine(min_score=min_score)

    # Match contacts to roles
    print("\n[3/6] Matching contacts to target roles...")
    matches = []
    match_stats = defaultdict(int)

    total_with_titles = df['Title / Role'].notna().sum()
    processed = 0

    for idx, row in df.iterrows():
        if pd.isna(row['Title / Role']) or not row['Title / Role']:
            continue

        processed += 1
        if processed % 1000 == 0:
            print(f"  Processed {processed:,} / {total_with_titles:,} contacts with titles...")

        # Match role
        matched_role, match_score, match_type, explanation = matcher.match_role(
            row['Title / Role'],
            row['Organization Type']
        )

        if matched_role:
            # Extract original confidence from Notes if available
            original_confidence = 50  # default
            if pd.notna(row.get('Notes')):
                conf_match = re.search(r'Confidence:\s*(\d+)', str(row['Notes']))
                if conf_match:
                    original_confidence = int(conf_match.group(1))

            # Calculate priority score
            # match_score (50%) + original_confidence (30%) + email_quality (20%)
            email_quality = 100 if pd.notna(row.get('Email Address')) else 0
            priority_score = (match_score * 0.5) + (original_confidence * 0.3) + (email_quality * 0.2)

            # Determine priority tier
            if priority_score >= 90:
                priority_tier = "High"
            elif priority_score >= 75:
                priority_tier = "Medium"
            else:
                priority_tier = "Low"

            # Create match record
            match_record = row.to_dict()
            match_record['Target Role'] = matched_role
            match_record['Match Score'] = int(match_score)
            match_record['Match Type'] = match_type
            match_record['Match Explanation'] = explanation
            match_record['Priority Score'] = round(priority_score, 1)
            match_record['Priority Tier'] = priority_tier

            matches.append(match_record)
            match_stats[matched_role] += 1

    print(f"\n  ✅ Matched {len(matches):,} contacts to target roles")

    # Create filtered dataframe
    print("\n[4/6] Creating filtered dataset...")
    df_filtered = pd.DataFrame(matches)

    # Remove duplicates (keep highest priority)
    print("  Removing duplicates (keeping highest priority match)...")
    df_filtered = df_filtered.sort_values('Priority Score', ascending=False)
    df_filtered = df_filtered.drop_duplicates(subset=['Email Address'], keep='first')

    print(f"  After deduplication: {len(df_filtered):,} unique contacts")

    # Print match statistics
    print("\n[5/6] Match statistics by role:")
    print("-" * 80)

    law_roles = [r for r in matcher.LAW_SCHOOL_ROLES.keys()]
    para_roles = [r for r in matcher.PARALEGAL_ROLES.keys()]

    print("\n  Law Schools:")
    law_total = 0
    for role in law_roles:
        count = len(df_filtered[(df_filtered['Target Role'] == role) & (df_filtered['Organization Type'] == 'Law School')])
        print(f"    {role:45s}: {count:4d} contacts")
        law_total += count
    print(f"    {'TOTAL LAW SCHOOLS':45s}: {law_total:4d} contacts")

    print("\n  Paralegal Programs:")
    para_total = 0
    for role in para_roles:
        count = len(df_filtered[(df_filtered['Target Role'] == role) & (df_filtered['Organization Type'] == 'Paralegal Program')])
        print(f"    {role:45s}: {count:4d} contacts")
        para_total += count
    print(f"    {'TOTAL PARALEGAL PROGRAMS':45s}: {para_total:4d} contacts")

    print(f"\n  {'GRAND TOTAL':45s}: {len(df_filtered):4d} contacts")

    # Priority tier breakdown
    print("\n  Priority Tier Breakdown:")
    for tier in ['High', 'Medium', 'Low']:
        count = (df_filtered['Priority Tier'] == tier).sum()
        pct = count / len(df_filtered) * 100 if len(df_filtered) > 0 else 0
        print(f"    {tier:10s}: {count:4d} contacts ({pct:5.1f}%)")

    # Generate Excel workbook
    print("\n[6/6] Generating Excel workbook with multiple sheets...")
    generate_excel_output(df_filtered, output_file, matcher)

    print(f"\n{'='*80}")
    print("✅ FILTERING COMPLETE!")
    print(f"{'='*80}")
    print(f"\nOutput file: {output_file}")
    print(f"Total filtered contacts: {len(df_filtered):,}")
    print(f"Reduction: {len(df):,} → {len(df_filtered):,} ({len(df_filtered)/len(df)*100:.1f}%)")

    return df_filtered


def generate_excel_output(df_filtered, output_file, matcher):
    """Generate multi-sheet Excel workbook with formatting"""

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define column order
    base_columns = [
        'Target Role', 'Match Score', 'Priority Score', 'Priority Tier',
        'Organization Type', 'Organization Name', 'State',
        'First Name', 'Last Name', 'Title / Role', 'Email Address',
        'Owner', 'Contact Source', 'Follow-Up Status',
        'Date 1st Email Sent', 'Date 2nd Follow-Up Sent',
        'Date 3rd Follow-Up Sent', 'Date 4th Follow-Up Sent',
        'Response Status?', 'Demo Date', 'Notes',
        'Match Type', 'Match Explanation'
    ]

    # Reorder columns
    df_ordered = df_filtered[base_columns].copy()

    # Sheet 1: All Priority Contacts
    print("  Creating sheet: All Priority Contacts...")
    ws_all = wb.create_sheet("All Priority Contacts")
    df_sorted = df_ordered.sort_values(['Priority Score', 'Match Score'], ascending=False)
    write_sheet_with_formatting(ws_all, df_sorted, "All Priority Contacts")

    # Sheet 2: Law Schools - Priority Contacts
    print("  Creating sheet: Law Schools...")
    ws_law = wb.create_sheet("Law Schools - Priority")
    df_law = df_ordered[df_ordered['Organization Type'] == 'Law School'].copy()
    df_law = df_law.sort_values(['Target Role', 'Priority Score'], ascending=[True, False])
    write_sheet_with_formatting(ws_law, df_law, "Law Schools")

    # Sheet 3: Paralegal Programs - Priority Contacts
    print("  Creating sheet: Paralegal Programs...")
    ws_para = wb.create_sheet("Paralegal - Priority")
    df_para = df_ordered[df_ordered['Organization Type'] == 'Paralegal Program'].copy()
    df_para = df_para.sort_values(['Target Role', 'Priority Score'], ascending=[True, False])
    write_sheet_with_formatting(ws_para, df_para, "Paralegal Programs")

    # Sheet 4-8: Individual role sheets (Law Schools)
    law_roles = list(matcher.LAW_SCHOOL_ROLES.keys())
    for role in law_roles[:3]:  # Top 3 roles
        safe_name = role[:25]  # Excel sheet name limit
        print(f"  Creating sheet: {safe_name}...")
        ws_role = wb.create_sheet(safe_name)
        df_role = df_ordered[(df_ordered['Target Role'] == role) &
                            (df_ordered['Organization Type'] == 'Law School')].copy()
        df_role = df_role.sort_values('Priority Score', ascending=False)
        write_sheet_with_formatting(ws_role, df_role, role)

    # Sheet 9: Statistics Summary
    print("  Creating sheet: Statistics Summary...")
    ws_stats = wb.create_sheet("Statistics Summary")
    create_statistics_sheet(ws_stats, df_filtered, matcher)

    # Save workbook
    wb.save(output_file)
    print(f"  ✅ Workbook saved: {output_file}")


def write_sheet_with_formatting(ws, df, sheet_name):
    """Write dataframe to sheet with professional formatting"""

    # Header formatting
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Priority tier colors
    tier_colors = {
        'High': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),      # Light green
        'Medium': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),    # Light yellow
        'Low': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")        # Light red
    }

    # Write headers
    for col_num, header in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data with priority tier coloring
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        priority_tier = row[df.columns.get_loc('Priority Tier')] if 'Priority Tier' in df.columns else None

        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)

            # Convert NaT to None
            if pd.isna(value):
                cell.value = None
            else:
                cell.value = value

            # Apply priority tier coloring to entire row
            if priority_tier and priority_tier in tier_colors:
                cell.fill = tier_colors[priority_tier]

    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"


def create_statistics_sheet(ws, df_filtered, matcher):
    """Create statistics summary dashboard"""

    # Title
    ws['A1'] = "ROLE-BASED FILTERING - STATISTICS SUMMARY"
    ws['A1'].font = Font(bold=True, size=14)

    row = 3

    # Overall statistics
    ws[f'A{row}'] = "OVERALL STATISTICS"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws[f'A{row}'] = "Total Filtered Contacts:"
    ws[f'B{row}'] = len(df_filtered)
    row += 1

    ws[f'A{row}'] = "Law School Contacts:"
    ws[f'B{row}'] = (df_filtered['Organization Type'] == 'Law School').sum()
    row += 1

    ws[f'A{row}'] = "Paralegal Program Contacts:"
    ws[f'B{row}'] = (df_filtered['Organization Type'] == 'Paralegal Program').sum()
    row += 2

    # Role breakdown
    ws[f'A{row}'] = "LAW SCHOOL ROLES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for role in matcher.LAW_SCHOOL_ROLES.keys():
        count = len(df_filtered[(df_filtered['Target Role'] == role) &
                               (df_filtered['Organization Type'] == 'Law School')])
        ws[f'A{row}'] = role
        ws[f'B{row}'] = count
        row += 1

    row += 1
    ws[f'A{row}'] = "PARALEGAL PROGRAM ROLES"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for role in matcher.PARALEGAL_ROLES.keys():
        count = len(df_filtered[(df_filtered['Target Role'] == role) &
                               (df_filtered['Organization Type'] == 'Paralegal Program')])
        ws[f'A{row}'] = role
        ws[f'B{row}'] = count
        row += 1

    row += 1
    ws[f'A{row}'] = "PRIORITY TIER BREAKDOWN"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    for tier in ['High', 'Medium', 'Low']:
        count = (df_filtered['Priority Tier'] == tier).sum()
        pct = count / len(df_filtered) * 100 if len(df_filtered) > 0 else 0
        ws[f'A{row}'] = tier
        ws[f'B{row}'] = count
        ws[f'C{row}'] = f"{pct:.1f}%"
        row += 1

    # Auto-size columns
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15


if __name__ == "__main__":
    # Configuration
    input_file = "output/contacts_BD_format.xlsx"
    output_file = "output/contacts_BD_format_FILTERED.xlsx"
    min_match_score = 65

    # Run filtering
    df_filtered = filter_contacts_by_roles(
        input_file=input_file,
        output_file=output_file,
        min_score=min_match_score
    )

    print("\n" + "=" * 80)
    print("FILTERING COMPLETE!")
    print("=" * 80)
    print(f"\nYou can now open: {output_file}")
    print("\nSheets included:")
    print("  1. All Priority Contacts - All filtered contacts sorted by priority")
    print("  2. Law Schools - Priority - Law school contacts only")
    print("  3. Paralegal - Priority - Paralegal program contacts only")
    print("  4-6. Individual role sheets (top 3 law school roles)")
    print("  7. Statistics Summary - Match statistics and breakdown")
