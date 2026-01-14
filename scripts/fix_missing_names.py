#!/usr/bin/env python3
"""
Fix Missing Names in Contact Database

Extracts first and last names from title field when they're missing.
Handles patterns like "FirstName LastName, Title" or "FirstName LastName Director of..."

Author: Generated for Legal Education Contact Scraper
Date: 2026-01-09
"""

import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


def extract_name_from_title(title):
    """
    Extract first and last name from title field

    Common patterns:
    - "Ashley Arrington Associate Director of..."
    - "Kevin Gerson, Law Library Director"
    - "Aysha Ames Director of Legal Writing aames4@..."
    - "Taylor, Abijah Assistant Dean for..."
    """
    if pd.isna(title) or not title:
        return None, None

    title = str(title).strip()

    # Pattern 1: "LastName, FirstName Title..." (last name first with comma)
    match = re.match(r'^([A-Z][a-zA-Z\-\']+),\s+([A-Z][a-zA-Z\-\']+)', title)
    if match:
        last_name = match.group(1)
        first_name = match.group(2)
        return first_name, last_name

    # Pattern 2: "FirstName LastName, Title..." (comma separator)
    match = re.match(r'^([A-Z][a-zA-Z\-\']+)\s+([A-Z][a-zA-Z\-\']+),', title)
    if match:
        first_name = match.group(1)
        last_name = match.group(2)
        return first_name, last_name

    # Pattern 3: "FirstName LastName Title..." (space before title word)
    # Look for pattern: FirstName LastName [Title words like Director, Dean, Professor, etc.]
    title_words = ['director', 'dean', 'professor', 'associate', 'assistant',
                   'coordinator', 'librarian', 'chair', 'head', 'manager',
                   'specialist', 'officer', 'counsel']

    for title_word in title_words:
        # Case insensitive search for title word
        pattern = r'^([A-Z][a-zA-Z\-\']+)\s+([A-Z][a-zA-Z\-\']+)\s+(?:' + title_word + ')'
        match = re.search(pattern, title, re.IGNORECASE)
        if match:
            first_name = match.group(1)
            last_name = match.group(2)
            # Validate names (not too long, not title words themselves)
            if (len(first_name) <= 20 and len(last_name) <= 25 and
                first_name.lower() not in title_words and
                last_name.lower() not in title_words):
                return first_name, last_name

    # Pattern 4: "FirstName LastName email@..." (before email)
    match = re.match(r'^([A-Z][a-zA-Z\-\']+)\s+([A-Z][a-zA-Z\-\']+)\s+[a-z0-9]', title)
    if match:
        first_name = match.group(1)
        last_name = match.group(2)
        if len(first_name) <= 20 and len(last_name) <= 25:
            return first_name, last_name

    # Pattern 5: First two capitalized words (conservative)
    words = title.split()
    if len(words) >= 2:
        first_word = words[0]
        second_word = words[1]

        # Must start with capital, contain only letters/hyphens/apostrophes
        if (re.match(r'^[A-Z][a-zA-Z\-\']+$', first_word) and
            re.match(r'^[A-Z][a-zA-Z\-\']+$', second_word) and
            len(first_word) <= 20 and len(second_word) <= 25 and
            first_word.lower() not in title_words and
            second_word.lower() not in title_words):
            return first_word, second_word

    return None, None


def fix_missing_names_in_file(input_file, output_file):
    """Fix missing names in Excel file"""

    print("=" * 80)
    print("FIXING MISSING NAMES IN CONTACT DATABASE")
    print("=" * 80)
    print(f"\nInput: {input_file}")
    print(f"Output: {output_file}")

    # Load all sheets
    xl_file = pd.ExcelFile(input_file)
    print(f"\nSheets found: {xl_file.sheet_names}")

    # Process each sheet
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name in xl_file.sheet_names:
            print(f"\n[Processing sheet: {sheet_name}]")

            df = pd.read_excel(input_file, sheet_name=sheet_name)

            if 'First Name' not in df.columns or 'Last Name' not in df.columns:
                print(f"  Skipping (no name columns)")
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                continue

            # Count missing names before
            missing_before = df['First Name'].isna().sum()
            print(f"  Missing names before: {missing_before:,} / {len(df):,}")

            # Fix missing names
            fixed_count = 0
            for idx, row in df.iterrows():
                if pd.isna(row['First Name']) or pd.isna(row['Last Name']):
                    # Try to extract from title
                    first, last = extract_name_from_title(row.get('Title / Role'))

                    if first and last:
                        df.at[idx, 'First Name'] = first
                        df.at[idx, 'Last Name'] = last
                        fixed_count += 1

                        if fixed_count <= 5:  # Show first 5 fixes
                            print(f"    Fixed: {first} {last} (from: {str(row.get('Title / Role'))[:60]}...)")

            # Count missing names after
            missing_after = df['First Name'].isna().sum()
            print(f"  Missing names after: {missing_after:,} / {len(df):,}")
            print(f"  Fixed: {fixed_count:,} names ({fixed_count/len(df)*100:.1f}%)")

            # Write to new file
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Copy formatting from original file
    print("\n[Copying formatting from original file...]")
    copy_formatting(input_file, output_file)

    print("\n" + "=" * 80)
    print("✅ FIXING COMPLETE!")
    print("=" * 80)


def copy_formatting(source_file, dest_file):
    """Copy cell formatting from source to destination"""

    try:
        # Load both workbooks
        wb_source = load_workbook(source_file)
        wb_dest = load_workbook(dest_file)

        # Copy formatting for each sheet
        for sheet_name in wb_source.sheetnames:
            if sheet_name not in wb_dest.sheetnames:
                continue

            ws_source = wb_source[sheet_name]
            ws_dest = wb_dest[sheet_name]

            # Copy column widths
            for col in ws_source.column_dimensions:
                if col in ws_dest.column_dimensions:
                    ws_dest.column_dimensions[col].width = ws_source.column_dimensions[col].width

            # Copy header formatting (first row)
            for col_idx in range(1, ws_source.max_column + 1):
                source_cell = ws_source.cell(1, col_idx)
                dest_cell = ws_dest.cell(1, col_idx)

                if source_cell.fill:
                    dest_cell.fill = source_cell.fill.copy()
                if source_cell.font:
                    dest_cell.font = source_cell.font.copy()
                if source_cell.alignment:
                    dest_cell.alignment = source_cell.alignment.copy()

            # Copy freeze panes
            if ws_source.freeze_panes:
                ws_dest.freeze_panes = ws_source.freeze_panes

        # Save
        wb_dest.save(dest_file)
        print("  Formatting copied successfully")

    except Exception as e:
        print(f"  Warning: Could not copy formatting: {e}")


if __name__ == "__main__":
    # Fix both BD format files

    print("\n" + "=" * 80)
    print("STEP 1: Fixing contacts_BD_format.xlsx")
    print("=" * 80)
    fix_missing_names_in_file(
        "output/contacts_BD_format.xlsx",
        "output/contacts_BD_format_FIXED.xlsx"
    )

    print("\n\n" + "=" * 80)
    print("STEP 2: Fixing contacts_BD_format_FILTERED.xlsx")
    print("=" * 80)
    fix_missing_names_in_file(
        "output/contacts_BD_format_FILTERED.xlsx",
        "output/contacts_BD_format_FILTERED_FIXED.xlsx"
    )

    print("\n\n" + "=" * 80)
    print("ALL FILES FIXED!")
    print("=" * 80)
    print("\nNew files created:")
    print("  1. output/contacts_BD_format_FIXED.xlsx")
    print("  2. output/contacts_BD_format_FILTERED_FIXED.xlsx")
    print("\nYou can now use these files with corrected names!")
