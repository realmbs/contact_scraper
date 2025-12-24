"""
Excel Output Module for Contact Scraper
Generates comprehensive multi-sheet Excel workbooks with formatting.

Sheet Structure (8 sheets):
    1. All Contacts - Every contact with all fields
    2. Law School Contacts - Filtered by program type
    3. Paralegal Program Contacts - Filtered by program type
    4. High Confidence - Score >= 75
    5. Medium Confidence - Score 50-74
    6. Needs Review - Score < 50
    7. Statistics Summary - Dashboard with metrics
    8. Scraping Log - Audit trail

Formatting Features:
    - Color coding by confidence level (green/yellow/red)
    - Frozen header rows
    - Auto-fit column widths
    - Data filters on headers
    - Conditional formatting for scores

Functions:
    - create_excel_workbook: Main Excel generator
    - add_all_contacts_sheet: Sheet 1 - All contacts
    - add_filtered_sheet: Sheets 2-6 - Filtered views
    - add_statistics_sheet: Sheet 7 - Stats dashboard
    - add_scraping_log_sheet: Sheet 8 - Audit trail
    - apply_formatting: Color coding, frozen panes, auto-fit, filters
"""

from typing import Dict, Any, List, Optional, Callable
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


# Color scheme for confidence levels
COLORS = {
    'high': 'C6EFCE',      # Light green
    'medium': 'FFEB9C',    # Light yellow
    'low': 'FFC7CE',       # Light red
    'header': '4472C4',    # Blue
    'header_text': 'FFFFFF'  # White
}


def apply_header_formatting(ws: Worksheet) -> None:
    """
    Apply formatting to header row (row 1).

    Args:
        ws: Worksheet to format
    """
    header_fill = PatternFill(start_color=COLORS['header'], end_color=COLORS['header'], fill_type='solid')
    header_font = Font(bold=True, color=COLORS['header_text'], size=11)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def apply_confidence_color_coding(ws: Worksheet, confidence_col_idx: int) -> None:
    """
    Apply color coding to rows based on confidence score.

    Args:
        ws: Worksheet to format
        confidence_col_idx: Column index for confidence_score (1-based)
    """
    high_fill = PatternFill(start_color=COLORS['high'], end_color=COLORS['high'], fill_type='solid')
    medium_fill = PatternFill(start_color=COLORS['medium'], end_color=COLORS['medium'], fill_type='solid')
    low_fill = PatternFill(start_color=COLORS['low'], end_color=COLORS['low'], fill_type='solid')

    # Skip header row (row 1)
    for row_idx in range(2, ws.max_row + 1):
        confidence_cell = ws.cell(row=row_idx, column=confidence_col_idx)

        try:
            score = float(confidence_cell.value) if confidence_cell.value is not None else 0

            # Determine fill color based on score
            if score >= 75:
                fill = high_fill
            elif score >= 50:
                fill = medium_fill
            else:
                fill = low_fill

            # Apply fill to entire row
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

        except (ValueError, TypeError):
            # Skip if confidence score is not a valid number
            continue


def auto_fit_columns(ws: Worksheet, max_width: int = 50) -> None:
    """
    Auto-fit column widths based on content.

    Args:
        ws: Worksheet to adjust
        max_width: Maximum column width (default: 50)
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def apply_data_filters(ws: Worksheet) -> None:
    """
    Add filter dropdowns to header row.

    Args:
        ws: Worksheet to add filters to
    """
    if ws.max_row > 1:  # Only add filters if there's data
        ws.auto_filter.ref = ws.dimensions


def freeze_header_row(ws: Worksheet) -> None:
    """
    Freeze the header row (row 1).

    Args:
        ws: Worksheet to freeze
    """
    ws.freeze_panes = 'A2'


def add_all_contacts_sheet(wb: Workbook, contacts_df: pd.DataFrame) -> None:
    """
    Add "All Contacts" sheet with every contact and all fields.

    Args:
        wb: Workbook to add sheet to
        contacts_df: DataFrame with all contacts
    """
    ws = wb.create_sheet(title="All Contacts")

    if contacts_df.empty:
        ws.append(["No contacts found"])
        logger.warning("No contacts to add to All Contacts sheet")
        return

    # Write data
    for row in dataframe_to_rows(contacts_df, index=False, header=True):
        ws.append(row)

    # Apply formatting
    apply_header_formatting(ws)
    freeze_header_row(ws)
    auto_fit_columns(ws)
    apply_data_filters(ws)

    # Apply color coding if confidence_score column exists
    if 'confidence_score' in contacts_df.columns:
        confidence_col_idx = contacts_df.columns.get_loc('confidence_score') + 1
        apply_confidence_color_coding(ws, confidence_col_idx)

    logger.info(f"Added All Contacts sheet: {len(contacts_df)} contacts")


def add_filtered_sheet(
    wb: Workbook,
    contacts_df: pd.DataFrame,
    sheet_name: str,
    filter_func: Callable[[pd.DataFrame], pd.DataFrame]
) -> None:
    """
    Add a filtered sheet based on a filter function.

    Args:
        wb: Workbook to add sheet to
        contacts_df: DataFrame with all contacts
        sheet_name: Name for the sheet
        filter_func: Function to filter the DataFrame
    """
    ws = wb.create_sheet(title=sheet_name)

    if contacts_df.empty:
        ws.append(["No contacts found"])
        logger.warning(f"No contacts to filter for {sheet_name} sheet")
        return

    # Apply filter
    filtered_df = filter_func(contacts_df)

    if filtered_df.empty:
        ws.append(["No contacts match filter criteria"])
        logger.info(f"{sheet_name} sheet: 0 contacts after filtering")
        return

    # Write data
    for row in dataframe_to_rows(filtered_df, index=False, header=True):
        ws.append(row)

    # Apply formatting
    apply_header_formatting(ws)
    freeze_header_row(ws)
    auto_fit_columns(ws)
    apply_data_filters(ws)

    # Apply color coding if confidence_score column exists
    if 'confidence_score' in filtered_df.columns:
        confidence_col_idx = filtered_df.columns.get_loc('confidence_score') + 1
        apply_confidence_color_coding(ws, confidence_col_idx)

    logger.info(f"Added {sheet_name} sheet: {len(filtered_df)} contacts")


def add_statistics_sheet(wb: Workbook, stats: Dict[str, Any]) -> None:
    """
    Add statistics summary sheet with metrics dashboard.

    Args:
        wb: Workbook to add sheet to
        stats: Dictionary with statistics from calculate_contact_statistics()
    """
    ws = wb.create_sheet(title="Statistics Summary")

    # Title
    ws['A1'] = "Contact Scraper Statistics Dashboard"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    current_row = 3

    # Summary statistics
    ws[f'A{current_row}'] = "SUMMARY"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    summary = stats.get('summary', {})
    ws.append(['Total Contacts', summary.get('total_contacts', 0)])
    ws.append(['Total Institutions', summary.get('total_institutions', 0)])
    ws.append(['Avg Contacts per Institution', summary.get('avg_contacts_per_institution', 0)])
    current_row += 4

    # Email quality metrics
    ws[f'A{current_row}'] = "EMAIL QUALITY"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    email = stats.get('email_quality', {})
    ws.append(['Contacts with Email', f"{email.get('with_email', 0)} ({email.get('email_coverage_pct', 0)}%)"])
    ws.append(['Valid Deliverable', f"{email.get('valid_deliverable', 0)} ({email.get('valid_pct', 0)}%)"])
    ws.append(['Catch-All Domains', f"{email.get('catch_all', 0)} ({email.get('catch_all_pct', 0)}%)"])
    ws.append(['Invalid Emails', f"{email.get('invalid', 0)} ({email.get('invalid_pct', 0)}%)"])
    ws.append(['Unknown Status', f"{email.get('unknown', 0)} ({email.get('unknown_pct', 0)}%)"])
    current_row += 6

    # Confidence distribution
    ws[f'A{current_row}'] = "CONFIDENCE DISTRIBUTION"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    confidence = stats.get('confidence_distribution', {})
    ws.append(['High (75-100)', f"{confidence.get('high', {}).get('count', 0)} ({confidence.get('high', {}).get('percentage', 0)}%)"])
    ws.append(['Medium (50-74)', f"{confidence.get('medium', {}).get('count', 0)} ({confidence.get('medium', {}).get('percentage', 0)}%)"])
    ws.append(['Low (0-49)', f"{confidence.get('low', {}).get('count', 0)} ({confidence.get('low', {}).get('percentage', 0)}%)"])
    ws.append(['Average Score', confidence.get('average_score', 0)])
    ws.append(['Median Score', confidence.get('median_score', 0)])
    current_row += 6

    # Scraping success rate
    ws[f'A{current_row}'] = "SCRAPING SUCCESS RATE"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    scraping = stats.get('scraping_success', {})
    ws.append(['Institutions Attempted', scraping.get('total_institutions_attempted', 0)])
    ws.append(['Successful Extractions', scraping.get('successful_extractions', 0)])
    ws.append(['Failed Extractions', scraping.get('failed_extractions', 0)])
    ws.append(['Success Rate', f"{scraping.get('success_rate_pct', 0)}%"])
    ws.append(['Avg Contacts per Institution', scraping.get('avg_contacts_per_institution', 0)])
    current_row += 6

    # Top roles
    ws[f'A{current_row}'] = "TOP ROLES"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    ws.append(['Role', 'Count'])
    top_roles = stats.get('top_roles', [])
    for role_data in top_roles[:10]:
        ws.append([role_data.get('role', ''), role_data.get('count', 0)])
    current_row += len(top_roles) + 2

    # State breakdown
    ws[f'A{current_row}'] = "CONTACTS BY STATE"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    ws.append(['State', 'Count'])
    state_breakdown = stats.get('by_state', {})
    for state, count in state_breakdown.items():
        ws.append([state, count])
    current_row += len(state_breakdown) + 2

    # Program type breakdown
    ws[f'A{current_row}'] = "CONTACTS BY PROGRAM TYPE"
    ws[f'A{current_row}'].font = Font(bold=True, size=12)
    current_row += 1

    ws.append(['Program Type', 'Count', 'Percentage'])
    program_breakdown = stats.get('by_program_type', {})
    for program, data in program_breakdown.items():
        ws.append([program, data.get('count', 0), f"{data.get('percentage', 0)}%"])

    # Auto-fit columns
    auto_fit_columns(ws)

    logger.info("Added Statistics Summary sheet")


def add_scraping_log_sheet(
    wb: Workbook,
    contacts_df: pd.DataFrame,
    targets_df: Optional[pd.DataFrame] = None
) -> None:
    """
    Add scraping log sheet with audit trail.

    Args:
        wb: Workbook to add sheet to
        contacts_df: DataFrame with contacts (to extract scraping metrics)
        targets_df: DataFrame with target institutions (optional)
    """
    ws = wb.create_sheet(title="Scraping Log")

    # Headers
    headers = ['Institution', 'State', 'Program Type', 'Status', 'Contacts Found', 'Source URL', 'Extracted At']
    ws.append(headers)

    if contacts_df.empty:
        ws.append(['No scraping data available'])
        logger.warning("No scraping data for log sheet")
        return

    # Group contacts by institution
    if 'institution_name' in contacts_df.columns:
        institution_groups = contacts_df.groupby('institution_name')

        for institution, group in institution_groups:
            state = group['state'].iloc[0] if 'state' in group.columns else 'N/A'
            program_type = group['program_type'].iloc[0] if 'program_type' in group.columns else 'N/A'
            contacts_found = len(group)
            source_url = group['source_url'].iloc[0] if 'source_url' in group.columns else 'N/A'
            extracted_at = group['extracted_at'].iloc[0] if 'extracted_at' in group.columns else 'N/A'

            ws.append([
                institution,
                state,
                program_type,
                'Success',
                contacts_found,
                source_url,
                extracted_at
            ])

    # Add failed institutions if targets_df provided
    if targets_df is not None and not targets_df.empty:
        if 'institution_name' in contacts_df.columns and 'name' in targets_df.columns:
            successful_institutions = set(contacts_df['institution_name'].unique())
            target_institutions = set(targets_df['name'].unique())
            failed_institutions = target_institutions - successful_institutions

            for institution in failed_institutions:
                target_row = targets_df[targets_df['name'] == institution].iloc[0]
                state = target_row.get('state', 'N/A')
                program_type = target_row.get('type', 'N/A')
                url = target_row.get('url', 'N/A')

                ws.append([
                    institution,
                    state,
                    program_type,
                    'Failed',
                    0,
                    url,
                    'N/A'
                ])

    # Apply formatting
    apply_header_formatting(ws)
    freeze_header_row(ws)
    auto_fit_columns(ws)
    apply_data_filters(ws)

    logger.info(f"Added Scraping Log sheet")


def create_excel_workbook(
    contacts_df: pd.DataFrame,
    stats: Dict[str, Any],
    output_path: str,
    targets_df: Optional[pd.DataFrame] = None
) -> bool:
    """
    Create comprehensive Excel workbook with all 8 sheets.

    Args:
        contacts_df: DataFrame with all contacts
        stats: Statistics dictionary from calculate_contact_statistics()
        output_path: Path to save Excel file
        targets_df: DataFrame with target institutions (optional, for scraping log)

    Returns:
        True if successful, False otherwise
    """
    try:
        logger.info(f"Creating Excel workbook: {output_path}")

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Sheet 1: All Contacts
        add_all_contacts_sheet(wb, contacts_df)

        # Sheet 2: Law School Contacts
        add_filtered_sheet(
            wb,
            contacts_df,
            "Law School Contacts",
            lambda df: df[df['program_type'] == 'Law School'] if 'program_type' in df.columns else pd.DataFrame()
        )

        # Sheet 3: Paralegal Program Contacts
        add_filtered_sheet(
            wb,
            contacts_df,
            "Paralegal Program Contacts",
            lambda df: df[df['program_type'] == 'Paralegal Program'] if 'program_type' in df.columns else pd.DataFrame()
        )

        # Sheet 4: High Confidence (>= 75)
        add_filtered_sheet(
            wb,
            contacts_df,
            "High Confidence",
            lambda df: df[df['confidence_score'] >= 75] if 'confidence_score' in df.columns else pd.DataFrame()
        )

        # Sheet 5: Medium Confidence (50-74)
        add_filtered_sheet(
            wb,
            contacts_df,
            "Medium Confidence",
            lambda df: df[(df['confidence_score'] >= 50) & (df['confidence_score'] < 75)] if 'confidence_score' in df.columns else pd.DataFrame()
        )

        # Sheet 6: Needs Review (< 50)
        add_filtered_sheet(
            wb,
            contacts_df,
            "Needs Review",
            lambda df: df[df['confidence_score'] < 50] if 'confidence_score' in df.columns else pd.DataFrame()
        )

        # Sheet 7: Statistics Summary
        add_statistics_sheet(wb, stats)

        # Sheet 8: Scraping Log
        add_scraping_log_sheet(wb, contacts_df, targets_df)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Excel workbook created successfully: {output_path}")

        return True

    except Exception as e:
        logger.error(f"Failed to create Excel workbook: {e}")
        return False
